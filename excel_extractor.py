import openpyxl
import pandas as pd
from typing import Dict, List, Any, Union, Optional
import io
import json
from datetime import datetime, date


class ExcelExtractor:
    """
    Extract Excel file content to a generic JSON format.
    Handles polymorphic Excel files with varying structures.
    """

    def __init__(self):
        self.date_formats = ['%Y-%m-%d', '%m/%d/%Y', '%d-%m-%Y', '%Y/%m/%d']

    def extract_to_json(self, file_content: bytes, filename: str) -> Dict[str, Any]:
        """
        Extract Excel file content to a generic JSON structure.

        Args:
            file_content: Excel file as bytes
            filename: Original filename

        Returns:
            Dictionary with extracted data in generic format
        """
        result = {
            'filename': filename,
            'extracted_at': datetime.now().isoformat(),
            'sheets': []
        }

        try:
            # Load workbook from bytes
            workbook = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_data = self._extract_sheet_data(sheet, sheet_name)
                result['sheets'].append(sheet_data)

            workbook.close()

        except Exception as e:
            # Fallback to pandas for older Excel formats
            try:
                result['sheets'] = self._extract_with_pandas(file_content)
            except Exception as pandas_error:
                raise Exception(f"Failed to extract Excel content: {str(e)}. Pandas fallback also failed: {str(pandas_error)}")

        return result

    def _extract_sheet_data(self, sheet, sheet_name: str) -> Dict[str, Any]:
        """
        Extract data from a single sheet.

        Args:
            sheet: openpyxl worksheet object
            sheet_name: Name of the sheet

        Returns:
            Dictionary with sheet data
        """
        sheet_result = {
            'sheet_name': sheet_name,
            'data_format': 'auto-detected',
            'rows': [],
            'metadata': {
                'total_rows': 0,
                'total_columns': 0,
                'has_header': False,
                'header_row': None
            }
        }

        # Get all rows as values
        all_rows = list(sheet.iter_rows(values_only=True))

        if not all_rows:
            return sheet_result

        # Detect if first row is a header (contains mostly strings)
        first_row = all_rows[0]
        header_detected = self._is_header_row(first_row)

        if header_detected:
            # Use first row as header
            headers = [self._clean_header(cell) for cell in first_row]
            sheet_result['metadata']['has_header'] = True
            sheet_result['metadata']['header_row'] = headers

            # Extract data rows as dictionaries
            for row_idx, row in enumerate(all_rows[1:], start=2):
                row_dict = {}
                for col_idx, (header, cell_value) in enumerate(zip(headers, row)):
                    if header:
                        row_dict[header] = self._serialize_value(cell_value)
                    else:
                        row_dict[f'column_{col_idx}'] = self._serialize_value(cell_value)

                # Only add non-empty rows
                if any(v is not None and v != '' for v in row_dict.values()):
                    row_dict['_row_number'] = row_idx
                    sheet_result['rows'].append(row_dict)

        else:
            # No header detected, use generic column names
            max_cols = max(len(row) for row in all_rows) if all_rows else 0

            for row_idx, row in enumerate(all_rows, start=1):
                row_dict = {}
                for col_idx, cell_value in enumerate(row):
                    row_dict[f'column_{col_idx}'] = self._serialize_value(cell_value)

                # Fill missing columns with None
                for col_idx in range(len(row), max_cols):
                    row_dict[f'column_{col_idx}'] = None

                # Only add non-empty rows
                if any(v is not None and v != '' for v in row_dict.values()):
                    row_dict['_row_number'] = row_idx
                    sheet_result['rows'].append(row_dict)

        sheet_result['metadata']['total_rows'] = len(sheet_result['rows'])
        sheet_result['metadata']['total_columns'] = len(sheet_result['rows'][0]) - 1 if sheet_result['rows'] else 0

        return sheet_result

    def _extract_with_pandas(self, file_content: bytes) -> List[Dict[str, Any]]:
        """
        Fallback extraction using pandas for older Excel formats.

        Args:
            file_content: Excel file as bytes

        Returns:
            List of sheet data dictionaries
        """
        sheets_data = []

        # Read all sheets
        excel_file = pd.ExcelFile(io.BytesIO(file_content))

        for sheet_name in excel_file.sheet_names:
            df = pd.read_excel(excel_file, sheet_name=sheet_name)

            # Convert DataFrame to records
            records = df.to_dict('records')

            # Serialize records
            serialized_records = []
            for idx, record in enumerate(records, start=2):  # Start at 2 assuming header at row 1
                serialized_record = {k: self._serialize_value(v) for k, v in record.items()}
                serialized_record['_row_number'] = idx
                serialized_records.append(serialized_record)

            sheet_data = {
                'sheet_name': sheet_name,
                'data_format': 'pandas-extracted',
                'rows': serialized_records,
                'metadata': {
                    'total_rows': len(serialized_records),
                    'total_columns': len(df.columns),
                    'has_header': True,
                    'header_row': df.columns.tolist()
                }
            }

            sheets_data.append(sheet_data)

        return sheets_data

    @staticmethod
    def _is_header_row(row: tuple) -> bool:
        """
        Detect if a row is likely a header row.

        Args:
            row: Tuple of cell values

        Returns:
            True if row appears to be a header
        """
        if not row:
            return False

        non_empty_cells = [cell for cell in row if cell is not None and cell != '']

        if not non_empty_cells:
            return False

        # Header detection: mostly strings, not too many numbers
        string_count = sum(1 for cell in non_empty_cells if isinstance(cell, str))
        number_count = sum(1 for cell in non_empty_cells if isinstance(cell, (int, float)))

        # If more than 70% are strings, likely a header
        return string_count > (len(non_empty_cells) * 0.7)

    @staticmethod
    def _clean_header(value: Any) -> str:
        """
        Clean and normalize header values.

        Args:
            value: Raw header value

        Returns:
            Cleaned header string
        """
        if value is None:
            return ''

        header = str(value).strip()

        # Replace spaces and special characters
        header = header.replace(' ', '_').replace('\n', '_')

        # Remove multiple underscores
        while '__' in header:
            header = header.replace('__', '_')

        return header

    def _serialize_value(self, value: Any) -> Union[str, int, float, bool, None]:
        """
        Serialize cell values to JSON-compatible types.

        Args:
            value: Cell value

        Returns:
            JSON-serializable value
        """
        if value is None or (isinstance(value, str) and value.strip() == ''):
            return None

        # Handle dates and datetimes
        if isinstance(value, (datetime, date)):
            return value.isoformat()

        # Handle pandas NaT and NaN
        if pd.isna(value):
            return None

        # Handle boolean
        if isinstance(value, bool):
            return value

        # Handle numbers
        if isinstance(value, (int, float)):
            # Check for NaN or Infinity
            if pd.isna(value):
                return None
            return value

        # Handle strings
        if isinstance(value, str):
            return value.strip()

        # Default: convert to string
        return str(value)

    def extract_summary(self, json_data: Dict[str, Any]) -> Dict[str, Any]:
        """
        Generate a summary of the extracted data.

        Args:
            json_data: Extracted JSON data

        Returns:
            Summary dictionary
        """
        summary = {
            'filename': json_data.get('filename'),
            'extracted_at': json_data.get('extracted_at'),
            'total_sheets': len(json_data.get('sheets', [])),
            'sheets_summary': []
        }

        for sheet in json_data.get('sheets', []):
            sheet_summary = {
                'sheet_name': sheet['sheet_name'],
                'total_rows': sheet['metadata']['total_rows'],
                'total_columns': sheet['metadata']['total_columns'],
                'has_header': sheet['metadata']['has_header'],
                'sample_data': sheet['rows'][:3] if sheet['rows'] else []
            }
            summary['sheets_summary'].append(sheet_summary)

        return summary

    def filter_rows_by_column(
        self,
        json_data: Dict[str, Any],
        column_name: str,
        filter_value: Any,
        sheet_name: Optional[str] = None
    ) -> List[Dict[str, Any]]:
        """
        Filter rows by column value.

        Args:
            json_data: Extracted JSON data
            column_name: Column to filter on
            filter_value: Value to match
            sheet_name: Optional sheet name to filter (if None, searches all sheets)

        Returns:
            List of matching rows
        """
        matching_rows = []

        for sheet in json_data.get('sheets', []):
            if sheet_name and sheet['sheet_name'] != sheet_name:
                continue

            for row in sheet['rows']:
                if column_name in row and row[column_name] == filter_value:
                    matching_rows.append({
                        'sheet': sheet['sheet_name'],
                        'row_number': row.get('_row_number'),
                        'data': row
                    })

        return matching_rows
