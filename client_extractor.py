"""
Client-centric Excel extractor with intelligent section detection.
Handles key-value pairs, tables, and complex headers with merged cells.
Enhanced with confidence scores and cell formatting metadata.
"""

import openpyxl
from openpyxl.cell import MergedCell
from openpyxl.styles import Font, PatternFill
from typing import Dict, List, Any, Optional, Tuple, Set
from datetime import datetime, date
import json
import hashlib


class ClientExtractor:
    """
    Extract client data from Excel files with intelligent section detection.
    """

    def __init__(self):
        self.min_key_value_rows = 2
        self.min_table_rows = 2
        self.cell_formatting_cache = {}  # Cache for cell formatting info

    def extract_client_data(
        self,
        file_path: str,
        client_info: Dict[str, Any]
    ) -> Dict[str, Any]:
        """
        Extract complete client data from Excel file.

        Args:
            file_path: Path to Excel file
            client_info: Client metadata (country, name, product, etc.)

        Returns:
            Complete client JSON document
        """
        result = {
            'client_id': client_info.get('client_id'),
            'client_name': client_info.get('client_name'),
            'country': client_info.get('country'),
            'product': client_info.get('product'),
            'file_info': {
                'file_path': file_path,
                'filename': client_info.get('filename'),
                'extracted_date': client_info.get('extracted_date').isoformat() if client_info.get('extracted_date') else None,
                'is_latest': client_info.get('is_latest', False),
                'form_variant': client_info.get('form_variant')
            },
            'sheets': [],
            'processing_metadata': {
                'processed_at': datetime.now().isoformat(),
                'status': 'success',
                'warnings': []
            }
        }

        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_data = self._extract_sheet_sections(sheet, sheet_name, result['processing_metadata']['warnings'])
                result['sheets'].append(sheet_data)

            workbook.close()

            # Generate pattern signature
            result['pattern_signature'] = self._generate_pattern_signature(result)

        except Exception as e:
            result['processing_metadata']['status'] = 'failed'
            result['processing_metadata']['error'] = str(e)

        return result

    def _extract_sheet_sections(
        self,
        sheet,
        sheet_name: str,
        warnings: List[str]
    ) -> Dict[str, Any]:
        """
        Extract all sections from a sheet.

        Args:
            sheet: openpyxl worksheet
            sheet_name: Name of sheet
            warnings: List to append warnings to

        Returns:
            Sheet data with detected sections
        """
        sheet_data = {
            'sheet_name': sheet_name,
            'sections': []
        }

        if sheet.max_row == 0:
            return sheet_data

        # Cache cell formatting for the sheet
        self._cache_cell_formatting(sheet)

        # Get merged cell ranges
        merged_ranges = self._get_merged_cell_ranges(sheet)

        # Get all cell data
        cell_matrix = self._build_cell_matrix(sheet, merged_ranges)

        # Identify section boundaries
        sections = self._identify_sections(cell_matrix, sheet.max_row, sheet.max_column)

        # Extract each section
        for idx, section_region in enumerate(sections):
            section = self._extract_section(
                cell_matrix,
                section_region,
                idx,
                merged_ranges,
                warnings,
                sheet
            )
            if section:
                sheet_data['sections'].append(section)

        return sheet_data

    def _cache_cell_formatting(self, sheet):
        """
        Cache cell formatting information for analysis.

        Args:
            sheet: openpyxl worksheet
        """
        self.cell_formatting_cache = {}

        for row in sheet.iter_rows():
            for cell in row:
                if not isinstance(cell, MergedCell):
                    formatting = {}

                    # Font properties
                    if cell.font:
                        formatting['bold'] = cell.font.bold or False
                        formatting['italic'] = cell.font.italic or False
                        formatting['font_size'] = cell.font.size
                        formatting['font_color'] = str(cell.font.color.rgb) if cell.font.color and hasattr(cell.font.color, 'rgb') else None

                    # Fill/background color
                    if cell.fill:
                        formatting['fill_type'] = cell.fill.patternType
                        if hasattr(cell.fill, 'fgColor') and cell.fill.fgColor:
                            formatting['fill_color'] = str(cell.fill.fgColor.rgb) if hasattr(cell.fill.fgColor, 'rgb') else None

                    # Border properties
                    if cell.border:
                        formatting['has_border'] = any([
                            cell.border.left and cell.border.left.style,
                            cell.border.right and cell.border.right.style,
                            cell.border.top and cell.border.top.style,
                            cell.border.bottom and cell.border.bottom.style
                        ])

                    # Alignment
                    if cell.alignment:
                        formatting['horizontal_align'] = cell.alignment.horizontal
                        formatting['vertical_align'] = cell.alignment.vertical

                    if formatting:
                        self.cell_formatting_cache[(cell.row, cell.column)] = formatting

    def _get_cell_formatting(self, row: int, col: int) -> Dict[str, Any]:
        """Get cached formatting for a cell."""
        return self.cell_formatting_cache.get((row, col), {})

    def _get_merged_cell_ranges(self, sheet) -> List[Dict[str, Any]]:
        """
        Get all merged cell ranges in the sheet.

        Returns:
            List of merged cell info
        """
        merged_ranges = []

        if hasattr(sheet, 'merged_cells'):
            for merged_range in sheet.merged_cells.ranges:
                merged_ranges.append({
                    'min_row': merged_range.min_row,
                    'max_row': merged_range.max_row,
                    'min_col': merged_range.min_col,
                    'max_col': merged_range.max_col,
                    'range': str(merged_range)
                })

        return merged_ranges

    def _build_cell_matrix(
        self,
        sheet,
        merged_ranges: List[Dict[str, Any]]
    ) -> Dict[Tuple[int, int], Any]:
        """
        Build cell matrix with merged cell propagation.

        Args:
            sheet: openpyxl worksheet
            merged_ranges: List of merged cell ranges

        Returns:
            Dictionary mapping (row, col) to cell value
        """
        cell_matrix = {}

        # First pass: get all non-merged cells
        for row in sheet.iter_rows():
            for cell in row:
                if not isinstance(cell, MergedCell):
                    cell_matrix[(cell.row, cell.column)] = cell.value

        # Second pass: propagate merged cell values
        for merged_range in merged_ranges:
            # Get value from top-left cell
            top_left_value = cell_matrix.get(
                (merged_range['min_row'], merged_range['min_col'])
            )

            # Propagate to all cells in merge
            for row in range(merged_range['min_row'], merged_range['max_row'] + 1):
                for col in range(merged_range['min_col'], merged_range['max_col'] + 1):
                    cell_matrix[(row, col)] = top_left_value

        return cell_matrix

    def _identify_sections(
        self,
        cell_matrix: Dict[Tuple[int, int], Any],
        max_row: int,
        max_col: int
    ) -> List[Dict[str, int]]:
        """
        Identify distinct sections in the sheet.

        Sections are separated by:
        - Empty rows (2+ consecutive)
        - Section headers (single bold row or text indicator)

        Args:
            cell_matrix: Cell value matrix
            max_row: Maximum row number
            max_col: Maximum column number

        Returns:
            List of section regions
        """
        sections = []
        current_section_start = None
        empty_row_count = 0

        for row in range(1, max_row + 1):
            # Check if row is empty
            row_has_data = any(
                self._is_non_empty(cell_matrix.get((row, col)))
                for col in range(1, max_col + 1)
            )

            if row_has_data:
                if current_section_start is None:
                    current_section_start = row
                empty_row_count = 0
            else:
                empty_row_count += 1

                # If we hit 2+ empty rows, consider it a section break
                if empty_row_count >= 2 and current_section_start is not None:
                    sections.append({
                        'start_row': current_section_start,
                        'end_row': row - empty_row_count,
                        'start_col': 1,
                        'end_col': max_col
                    })
                    current_section_start = None

        # Add final section if exists
        if current_section_start is not None:
            sections.append({
                'start_row': current_section_start,
                'end_row': max_row,
                'start_col': 1,
                'end_col': max_col
            })

        return sections

    def _extract_section(
        self,
        cell_matrix: Dict[Tuple[int, int], Any],
        region: Dict[str, int],
        section_idx: int,
        merged_ranges: List[Dict[str, Any]],
        warnings: List[str],
        sheet
    ) -> Optional[Dict[str, Any]]:
        """
        Extract a single section and determine its type.

        Args:
            cell_matrix: Cell value matrix
            region: Section region
            section_idx: Section index
            merged_ranges: Merged cell ranges
            warnings: List to append warnings
            sheet: openpyxl worksheet

        Returns:
            Section data dictionary with confidence score
        """
        start_row = region['start_row']
        end_row = region['end_row']
        start_col = region['start_col']
        end_col = region['end_col']

        # Extract raw data from region
        raw_data = []
        for row in range(start_row, end_row + 1):
            row_data = []
            for col in range(start_col, end_col + 1):
                value = cell_matrix.get((row, col))
                row_data.append(value)
            raw_data.append(row_data)

        # Detect section type with confidence
        section_type, section_header, confidence = self._detect_section_type_with_confidence(
            raw_data, merged_ranges, region
        )

        # Extract based on type
        section_data = None
        if section_type == 'key_value':
            section_data = self._extract_key_value_section(raw_data, region, section_idx, section_header)
        elif section_type == 'table':
            section_data = self._extract_table_section(raw_data, region, section_idx, section_header, warnings)
        elif section_type == 'complex_header':
            section_data = self._extract_complex_header_section(
                raw_data, region, section_idx, section_header, merged_ranges, warnings
            )
        else:
            # Unknown type - store as raw
            section_data = self._extract_raw_section(raw_data, region, section_idx)

        # Add confidence score and metadata
        if section_data:
            section_data['detection_confidence'] = confidence
            section_data['cell_coordinates'] = {
                'start_row': start_row,
                'end_row': end_row,
                'start_col': start_col,
                'end_col': end_col
            }

            # Add formatting hints for header row
            if section_type in ['table', 'complex_header']:
                header_formatting = self._get_cell_formatting(start_row, start_col)
                if header_formatting:
                    section_data['header_formatting'] = header_formatting

        return section_data

    def _detect_section_type_with_confidence(
        self,
        raw_data: List[List[Any]],
        merged_ranges: List[Dict[str, Any]],
        region: Dict[str, int]
    ) -> Tuple[str, Optional[str], float]:
        """
        Detect section type with confidence score.

        Returns:
            (section_type, section_header, confidence_score)
        """
        if not raw_data or len(raw_data) < 2:
            return ('raw', None, 1.0)

        confidence_scores = {
            'key_value': 0.0,
            'table': 0.0,
            'complex_header': 0.0,
            'raw': 0.5  # default baseline
        }

        # Check for section header (first row is single cell or title)
        section_header = None
        first_row = raw_data[0]
        first_row_non_empty = [v for v in first_row if self._is_non_empty(v)]

        # Check if first row is bold (indicates header)
        first_row_formatting = self._get_cell_formatting(region['start_row'], region['start_col'])
        is_first_row_bold = first_row_formatting.get('bold', False)

        if len(first_row_non_empty) == 1 and isinstance(first_row_non_empty[0], str):
            # Likely a section header
            section_header = str(first_row_non_empty[0]).strip()
            data_rows = raw_data[1:]
            if is_first_row_bold:
                confidence_scores['table'] += 0.2
                confidence_scores['complex_header'] += 0.2
        else:
            data_rows = raw_data

        if len(data_rows) < 2:
            return ('raw', section_header, 1.0)

        # Check for merged cells in header area (indicates complex header)
        has_merged_in_header = any(
            mr['min_row'] >= region['start_row'] and
            mr['max_row'] <= region['start_row'] + 3 and
            (mr['max_row'] - mr['min_row'] > 0 or mr['max_col'] - mr['min_col'] > 0)
            for mr in merged_ranges
        )

        if has_merged_in_header:
            confidence_scores['complex_header'] += 0.8

        # Check for key-value pattern (2 columns, mostly strings in col 1)
        non_empty_cols = self._count_non_empty_columns(data_rows)

        if non_empty_cols <= 2:
            confidence_scores['key_value'] += 0.3

            # Check if first column has string labels
            first_col_strings = sum(
                1 for row in data_rows
                if row and isinstance(row[0], str) and self._is_non_empty(row[0])
            )

            first_col_ratio = first_col_strings / len(data_rows) if len(data_rows) > 0 else 0
            confidence_scores['key_value'] += first_col_ratio * 0.7

        # Check for table indicators
        if non_empty_cols > 2:
            confidence_scores['table'] += 0.5

        # Check if first data row has uniform types (header-like)
        if data_rows:
            first_data_row = data_rows[0]
            string_count = sum(1 for v in first_data_row if isinstance(v, str) and self._is_non_empty(v))
            if string_count >= len([v for v in first_data_row if self._is_non_empty(v)]) * 0.8:
                confidence_scores['table'] += 0.3

        # Determine winner
        section_type = max(confidence_scores, key=confidence_scores.get)
        confidence = confidence_scores[section_type]

        # Normalize confidence to 0-1 range
        confidence = min(1.0, max(0.0, confidence))

        return (section_type, section_header, confidence)

    def _detect_section_type(
        self,
        raw_data: List[List[Any]],
        merged_ranges: List[Dict[str, Any]],
        region: Dict[str, int]
    ) -> Tuple[str, Optional[str]]:
        """
        Detect section type: key_value, table, or complex_header.
        Legacy method - calls confidence-based version.

        Returns:
            (section_type, section_header)
        """
        section_type, section_header, _ = self._detect_section_type_with_confidence(
            raw_data, merged_ranges, region
        )
        return (section_type, section_header)

    def _extract_key_value_section(
        self,
        raw_data: List[List[Any]],
        region: Dict[str, int],
        section_idx: int,
        section_header: Optional[str]
    ) -> Dict[str, Any]:
        """Extract key-value section."""
        data = {}

        for row in raw_data:
            if len(row) >= 2:
                key = row[0]
                value = row[1]

                if self._is_non_empty(key):
                    clean_key = self._clean_key(key)
                    data[clean_key] = self._serialize_value(value)

        return {
            'section_id': f'section_{section_idx}',
            'section_type': 'key_value',
            'section_header': section_header,
            'region': region,
            'data': data
        }

    def _extract_table_section(
        self,
        raw_data: List[List[Any]],
        region: Dict[str, int],
        section_idx: int,
        section_header: Optional[str],
        warnings: List[str]
    ) -> Dict[str, Any]:
        """Extract standard table section."""
        if not raw_data:
            return None

        # Assume first row is header
        header_row = raw_data[0]
        headers = [self._clean_key(h) if h else f'column_{i}' for i, h in enumerate(header_row)]

        # Extract data rows
        data_rows = []
        for row_idx, row in enumerate(raw_data[1:], start=1):
            row_dict = {}
            for col_idx, (header, value) in enumerate(zip(headers, row)):
                row_dict[header] = self._serialize_value(value)

            # Only add non-empty rows
            if any(v is not None and v != '' for v in row_dict.values()):
                row_dict['_row_number'] = region['start_row'] + row_idx
                data_rows.append(row_dict)

        return {
            'section_id': f'section_{section_idx}',
            'section_type': 'table',
            'section_header': section_header,
            'region': region,
            'headers': headers,
            'data': data_rows
        }

    def _extract_complex_header_section(
        self,
        raw_data: List[List[Any]],
        region: Dict[str, int],
        section_idx: int,
        section_header: Optional[str],
        merged_ranges: List[Dict[str, Any]],
        warnings: List[str]
    ) -> Dict[str, Any]:
        """Extract section with complex multi-level headers."""
        if len(raw_data) < 3:
            # Not enough rows for complex header
            return self._extract_table_section(raw_data, region, section_idx, section_header, warnings)

        # Try to detect header rows (usually first 2-3 rows)
        # Look for merged cells
        header_rows = min(3, len(raw_data) - 1)

        # Build hierarchical column names
        final_headers = []
        num_cols = len(raw_data[0])

        for col_idx in range(num_cols):
            header_parts = []

            for header_row_idx in range(header_rows):
                if header_row_idx < len(raw_data):
                    value = raw_data[header_row_idx][col_idx]
                    if self._is_non_empty(value):
                        clean_value = self._clean_key(value)
                        if clean_value not in header_parts:
                            header_parts.append(clean_value)

            if header_parts:
                final_header = '_'.join(header_parts)
            else:
                final_header = f'column_{col_idx}'

            final_headers.append(final_header)

        # Extract data rows (skip header rows)
        data_rows = []
        for row_idx, row in enumerate(raw_data[header_rows:], start=header_rows):
            row_dict = {}
            for col_idx, (header, value) in enumerate(zip(final_headers, row)):
                row_dict[header] = self._serialize_value(value)

            if any(v is not None and v != '' for v in row_dict.values()):
                row_dict['_row_number'] = region['start_row'] + row_idx
                data_rows.append(row_dict)

        return {
            'section_id': f'section_{section_idx}',
            'section_type': 'complex_header',
            'section_header': section_header,
            'region': region,
            'header_structure': {
                'levels': header_rows,
                'final_columns': final_headers
            },
            'data': data_rows
        }

    def _extract_raw_section(
        self,
        raw_data: List[List[Any]],
        region: Dict[str, int],
        section_idx: int
    ) -> Dict[str, Any]:
        """Extract section as raw data (fallback)."""
        serialized_data = [
            [self._serialize_value(v) for v in row]
            for row in raw_data
        ]

        return {
            'section_id': f'section_{section_idx}',
            'section_type': 'raw',
            'section_header': None,
            'region': region,
            'data': serialized_data
        }

    def _count_non_empty_columns(self, rows: List[List[Any]]) -> int:
        """Count columns with at least one non-empty value."""
        if not rows:
            return 0

        max_cols = max(len(row) for row in rows)
        non_empty_cols = 0

        for col_idx in range(max_cols):
            has_data = any(
                col_idx < len(row) and self._is_non_empty(row[col_idx])
                for row in rows
            )
            if has_data:
                non_empty_cols += 1

        return non_empty_cols

    def _generate_pattern_signature(self, client_data: Dict[str, Any]) -> str:
        """
        Generate pattern signature for the client structure.

        Based on:
        - Sheet names
        - Section types and headers
        - Key field names
        """
        signature_parts = []

        # Sheet names
        sheet_names = [s['sheet_name'] for s in client_data.get('sheets', [])]
        signature_parts.append('sheets:' + '|'.join(sorted(sheet_names)))

        # Section structure
        for sheet in client_data.get('sheets', []):
            for section in sheet.get('sections', []):
                section_type = section.get('section_type')
                section_header = section.get('section_header', '')

                signature_parts.append(f'section:{section_type}:{section_header}')

                # Add field names
                if section_type == 'key_value':
                    keys = sorted(section.get('data', {}).keys())
                    signature_parts.append('keys:' + '|'.join(keys[:10]))  # First 10 keys
                elif section_type in ['table', 'complex_header']:
                    headers = section.get('headers', section.get('header_structure', {}).get('final_columns', []))
                    signature_parts.append('headers:' + '|'.join(sorted(headers[:10])))

        signature_str = '||'.join(signature_parts)
        return hashlib.md5(signature_str.encode()).hexdigest()

    @staticmethod
    def _is_non_empty(value: Any) -> bool:
        """Check if value is non-empty."""
        if value is None:
            return False
        if isinstance(value, str) and not value.strip():
            return False
        return True

    @staticmethod
    def _clean_key(value: Any) -> str:
        """Clean and normalize key/header values."""
        if value is None:
            return ''

        key = str(value).strip()
        key = key.replace(' ', '_').replace('\n', '_').replace('\t', '_')

        # Remove special characters
        key = ''.join(c if c.isalnum() or c == '_' else '_' for c in key)

        # Remove multiple underscores
        while '__' in key:
            key = key.replace('__', '_')

        # Remove leading/trailing underscores
        key = key.strip('_')

        return key

    def _serialize_value(self, value: Any) -> Any:
        """Serialize cell values to JSON-compatible types."""
        if value is None or (isinstance(value, str) and value.strip() == ''):
            return None

        if isinstance(value, (datetime, date)):
            return value.isoformat()

        if isinstance(value, bool):
            return value

        if isinstance(value, (int, float)):
            return value

        if isinstance(value, str):
            return value.strip()

        return str(value)
