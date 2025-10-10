"""
Enhanced Excel parser to identify table-like structures within sheets.
Handles polymorphic data structures and multiple tables per sheet.
"""

import openpyxl
from openpyxl.cell import MergedCell
import pandas as pd
from typing import Dict, List, Any, Tuple, Optional
import io
import json
from datetime import datetime, date
import hashlib


class TableExtractor:
    """
    Advanced Excel parser that identifies and extracts table structures.
    Handles polymorphic Excel files with varying layouts.
    """

    def __init__(self):
        self.min_table_rows = 2  # Minimum rows to consider as a table
        self.min_table_cols = 2  # Minimum columns to consider as a table

    def extract_tables_from_file(self, file_path: str, filename: str) -> Dict[str, Any]:
        """
        Extract all table structures from an Excel file.

        Args:
            file_path: Path to Excel file
            filename: Original filename

        Returns:
            Dictionary with extracted tables and metadata
        """
        result = {
            'filename': filename,
            'file_path': file_path,
            'extracted_at': datetime.now().isoformat(),
            'sheets': []
        }

        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_data = self._extract_sheet_tables(sheet, sheet_name)
                result['sheets'].append(sheet_data)

            workbook.close()

        except Exception as e:
            result['error'] = str(e)
            result['extraction_method'] = 'failed'

        return result

    def _extract_sheet_tables(self, sheet, sheet_name: str) -> Dict[str, Any]:
        """
        Extract all tables from a single sheet.

        Args:
            sheet: openpyxl worksheet object
            sheet_name: Name of the sheet

        Returns:
            Dictionary with sheet data including detected tables
        """
        sheet_result = {
            'sheet_name': sheet_name,
            'tables': [],
            'metadata': {
                'total_rows': sheet.max_row,
                'total_columns': sheet.max_column,
                'merged_cells': len(sheet.merged_cells.ranges) if hasattr(sheet, 'merged_cells') else 0
            }
        }

        # Get all cell data with formatting info
        cell_data = self._get_all_cells(sheet)

        # Identify table regions
        table_regions = self._identify_table_regions(cell_data, sheet.max_row, sheet.max_column)

        # Extract each table
        for idx, region in enumerate(table_regions):
            table = self._extract_table_from_region(sheet, region, idx)
            if table:
                sheet_result['tables'].append(table)

        # If no tables detected, treat entire sheet as one table
        if not sheet_result['tables'] and sheet.max_row > 0:
            full_sheet_table = self._extract_full_sheet_as_table(sheet)
            if full_sheet_table:
                sheet_result['tables'].append(full_sheet_table)

        sheet_result['metadata']['tables_detected'] = len(sheet_result['tables'])

        return sheet_result

    def _get_all_cells(self, sheet) -> Dict[Tuple[int, int], Dict[str, Any]]:
        """
        Get all cell data with metadata.

        Returns:
            Dictionary mapping (row, col) to cell info
        """
        cell_data = {}

        for row in sheet.iter_rows():
            for cell in row:
                if not isinstance(cell, MergedCell):
                    cell_info = {
                        'value': cell.value,
                        'is_empty': cell.value is None or (isinstance(cell.value, str) and not cell.value.strip()),
                        'data_type': type(cell.value).__name__,
                        'has_style': cell.has_style if hasattr(cell, 'has_style') else False
                    }
                    cell_data[(cell.row, cell.column)] = cell_info

        return cell_data

    def _identify_table_regions(self, cell_data: Dict, max_row: int, max_col: int) -> List[Dict[str, int]]:
        """
        Identify distinct table regions within the sheet.

        Args:
            cell_data: Cell data dictionary
            max_row: Maximum row number
            max_col: Maximum column number

        Returns:
            List of table regions with boundaries
        """
        regions = []

        # Find dense data regions (areas with high cell density)
        # This helps separate multiple tables on the same sheet

        # Simple approach: look for blocks separated by empty rows/columns
        visited_rows = set()

        for start_row in range(1, max_row + 1):
            if start_row in visited_rows:
                continue

            # Check if this row has data
            row_has_data = any(
                not cell_data.get((start_row, col), {}).get('is_empty', True)
                for col in range(1, max_col + 1)
            )

            if not row_has_data:
                continue

            # Find the end of this table region
            end_row = start_row
            empty_row_count = 0

            for r in range(start_row + 1, max_row + 1):
                row_has_data = any(
                    not cell_data.get((r, col), {}).get('is_empty', True)
                    for col in range(1, max_col + 1)
                )

                if row_has_data:
                    end_row = r
                    empty_row_count = 0
                else:
                    empty_row_count += 1
                    # If we hit 2+ consecutive empty rows, consider it a break
                    if empty_row_count >= 2:
                        break

            # Find column boundaries for this region
            occupied_cols = set()
            for r in range(start_row, end_row + 1):
                for c in range(1, max_col + 1):
                    if not cell_data.get((r, c), {}).get('is_empty', True):
                        occupied_cols.add(c)

            if occupied_cols:
                start_col = min(occupied_cols)
                end_col = max(occupied_cols)

                if end_row - start_row + 1 >= self.min_table_rows and end_col - start_col + 1 >= self.min_table_cols:
                    regions.append({
                        'start_row': start_row,
                        'end_row': end_row,
                        'start_col': start_col,
                        'end_col': end_col
                    })

                    # Mark rows as visited
                    for r in range(start_row, end_row + 1):
                        visited_rows.add(r)

        return regions

    def _extract_table_from_region(self, sheet, region: Dict[str, int], table_idx: int) -> Optional[Dict[str, Any]]:
        """
        Extract a table from a specific region.

        Args:
            sheet: openpyxl worksheet
            region: Region boundaries
            table_idx: Index of this table in the sheet

        Returns:
            Extracted table data
        """
        start_row = region['start_row']
        end_row = region['end_row']
        start_col = region['start_col']
        end_col = region['end_col']

        # Extract values from region
        rows_data = []
        for row_idx in range(start_row, end_row + 1):
            row_values = []
            for col_idx in range(start_col, end_col + 1):
                cell = sheet.cell(row_idx, col_idx)
                row_values.append(cell.value)
            rows_data.append(row_values)

        if not rows_data:
            return None

        # Detect header row
        header_row_idx = self._detect_header_row(rows_data)

        table_data = {
            'table_id': f"table_{table_idx}",
            'region': region,
            'header_detected': header_row_idx is not None,
            'header_row_index': header_row_idx,
            'rows': [],
            'metadata': {
                'total_rows': len(rows_data),
                'total_columns': end_col - start_col + 1,
                'data_start_row': start_row,
                'data_end_row': end_row,
                'pattern_signature': None
            }
        }

        # Extract data based on header detection
        if header_row_idx is not None:
            headers = [self._clean_header(h) for h in rows_data[header_row_idx]]

            for data_row_idx in range(header_row_idx + 1, len(rows_data)):
                row_dict = {}
                for col_idx, (header, value) in enumerate(zip(headers, rows_data[data_row_idx])):
                    key = header if header else f'column_{col_idx}'
                    row_dict[key] = self._serialize_value(value)

                # Only add non-empty rows
                if any(v is not None and v != '' for v in row_dict.values()):
                    row_dict['_row_number'] = start_row + data_row_idx
                    table_data['rows'].append(row_dict)

            table_data['metadata']['headers'] = headers
        else:
            # No header detected - use generic column names
            num_cols = end_col - start_col + 1

            for row_idx, row_values in enumerate(rows_data):
                row_dict = {}
                for col_idx, value in enumerate(row_values):
                    row_dict[f'column_{col_idx}'] = self._serialize_value(value)

                if any(v is not None and v != '' for v in row_dict.values()):
                    row_dict['_row_number'] = start_row + row_idx
                    table_data['rows'].append(row_dict)

        # Generate pattern signature for this table
        table_data['metadata']['pattern_signature'] = self._generate_pattern_signature(table_data)

        return table_data

    def _extract_full_sheet_as_table(self, sheet) -> Optional[Dict[str, Any]]:
        """
        Extract entire sheet as a single table (fallback).
        """
        all_rows = list(sheet.iter_rows(values_only=True))

        if not all_rows:
            return None

        # Detect header
        header_row_idx = self._detect_header_row(all_rows)

        table_data = {
            'table_id': 'table_0',
            'region': {
                'start_row': 1,
                'end_row': sheet.max_row,
                'start_col': 1,
                'end_col': sheet.max_column
            },
            'header_detected': header_row_idx is not None,
            'header_row_index': header_row_idx,
            'rows': [],
            'metadata': {
                'total_rows': len(all_rows),
                'total_columns': sheet.max_column,
                'is_full_sheet': True
            }
        }

        if header_row_idx is not None:
            headers = [self._clean_header(h) for h in all_rows[header_row_idx]]

            for row_idx in range(header_row_idx + 1, len(all_rows)):
                row_dict = {}
                for col_idx, (header, value) in enumerate(zip(headers, all_rows[row_idx])):
                    key = header if header else f'column_{col_idx}'
                    row_dict[key] = self._serialize_value(value)

                if any(v is not None and v != '' for v in row_dict.values()):
                    row_dict['_row_number'] = row_idx + 1
                    table_data['rows'].append(row_dict)

            table_data['metadata']['headers'] = headers
        else:
            for row_idx, row_values in enumerate(all_rows):
                row_dict = {}
                for col_idx, value in enumerate(row_values):
                    row_dict[f'column_{col_idx}'] = self._serialize_value(value)

                if any(v is not None and v != '' for v in row_dict.values()):
                    row_dict['_row_number'] = row_idx + 1
                    table_data['rows'].append(row_dict)

        table_data['metadata']['pattern_signature'] = self._generate_pattern_signature(table_data)

        return table_data

    def _detect_header_row(self, rows_data: List[List[Any]]) -> Optional[int]:
        """
        Detect which row is the header row.

        Returns:
            Index of header row or None if not detected
        """
        if not rows_data:
            return None

        # Check first few rows for header characteristics
        for idx in range(min(5, len(rows_data))):
            row = rows_data[idx]

            # Skip completely empty rows
            non_empty = [cell for cell in row if cell is not None and (not isinstance(cell, str) or cell.strip())]
            if not non_empty:
                continue

            # Header characteristics:
            # 1. Mostly strings
            # 2. Not too many numbers
            # 3. Values are relatively unique
            string_count = sum(1 for cell in non_empty if isinstance(cell, str))

            if string_count >= len(non_empty) * 0.7:  # 70% strings
                return idx

        return None

    def _generate_pattern_signature(self, table_data: Dict[str, Any]) -> str:
        """
        Generate a signature/hash for the table pattern.
        This helps identify similar table structures across files.

        Args:
            table_data: Extracted table data

        Returns:
            Pattern signature hash
        """
        # Use headers and data types to create signature
        signature_parts = []

        if table_data.get('metadata', {}).get('headers'):
            # Use header names
            headers = table_data['metadata']['headers']
            signature_parts.extend(headers)
        elif table_data['rows']:
            # Use column names from first row
            first_row = table_data['rows'][0]
            signature_parts.extend(sorted([k for k in first_row.keys() if not k.startswith('_')]))

        # Add row count range (bucketed)
        row_count = len(table_data['rows'])
        if row_count < 10:
            signature_parts.append('rows_lt_10')
        elif row_count < 50:
            signature_parts.append('rows_lt_50')
        elif row_count < 100:
            signature_parts.append('rows_lt_100')
        else:
            signature_parts.append('rows_gte_100')

        # Create hash
        signature_str = '|'.join(str(p) for p in signature_parts)
        return hashlib.md5(signature_str.encode()).hexdigest()

    @staticmethod
    def _clean_header(value: Any) -> str:
        """Clean and normalize header values."""
        if value is None:
            return ''

        header = str(value).strip()
        header = header.replace(' ', '_').replace('\n', '_').replace('\t', '_')

        # Remove multiple underscores
        while '__' in header:
            header = header.replace('__', '_')

        # Remove leading/trailing underscores
        header = header.strip('_')

        return header

    def _serialize_value(self, value: Any) -> Any:
        """Serialize cell values to JSON-compatible types."""
        if value is None or (isinstance(value, str) and value.strip() == ''):
            return None

        if isinstance(value, (datetime, date)):
            return value.isoformat()

        if pd.isna(value):
            return None

        if isinstance(value, bool):
            return value

        if isinstance(value, (int, float)):
            if pd.isna(value):
                return None
            return value

        if isinstance(value, str):
            return value.strip()

        return str(value)
